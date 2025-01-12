from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework.parsers import MultiPartParser
from django.http import FileResponse
from ..utils.services import StudentExcelService #, BillingExcelService
from api.models import Student, AcadTermBilling, Grade, Enrollment, BillingList, Receipt
from ..utils.validators import FileValidator
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.drawing.image import Image
from django.conf import settings
from django.db.models import Sum
from ..serializers import StudentSerializer, EnrollmentSerializer, ReceiptSerializer
from ..utils.filterer import QuerysetFilter
import os
from django.db import transaction
import pandas as pd

class StudentExcelAPI(APIView):
    parser_classes = [MultiPartParser]

    def post(self, request):
        file = request.FILES.get("file")
        if not file:
            return Response({"error": "No file uploaded."}, status=400)

        if not FileValidator.is_valid_excel(file):
            return Response({"error": "Invalid file format. Please upload an Excel file."}, status=400)

        try:
            service = StudentExcelService()
            data = service.read(file)
            result = service.process(data)
            return Response({"message": result}, status=200)
        except Exception as e:
            return Response({"error": str(e)}, status=500)


# class BillingExcelAPI(APIView):
#     parser_classes = [MultiPartParser]

#     def post(self, request):
#         file = request.FILES.get("file")
#         if not file:
#             return Response({"error": "No file uploaded."}, status=400)

#         if not FileValidator.is_valid_excel(file):
#             return Response({"error": "Invalid file format. Please upload an Excel file."}, status=400)

#         try:
#             service = BillingExcelService()
#             data = service.read(file)
#             result = service.process(data)
#             return Response({"message": result}, status=200)
#         except Exception as e:
#             return Response({"error": str(e)}, status=500)

class GenerateCORAPI(APIView):
    # API endpoint for generating a Certificate of Registration (COR).

    def get(self, request, student_id):
        try:
            # Fetch the student data
            student = Student.objects.get(id=student_id)
            student_data = StudentSerializer(student).data

            # Fetch enrollments
            enrollments = Enrollment.objects.filter(student=student, school_year=student_data['academic_year'])
            enrollments_data = EnrollmentSerializer(enrollments, many=True).data

            # Fetch billing data
            billing_list = BillingList.objects.all()
            joined_data = []
            for billing in billing_list:
                acad_term_billing = AcadTermBilling.objects.filter(
                    billing=billing,
                    year_level=student_data['year_level'],
                    semester=student_data['semester']
                ).first()

                joined_data.append({
                    "billing_list": {
                        "name": billing.name,
                        "category": billing.category,
                        "price": acad_term_billing.price if acad_term_billing else None,
                        "year_level": acad_term_billing.year_level if acad_term_billing else None,
                        "semester": acad_term_billing.semester if acad_term_billing else None
                    }
                })

            # Calculate total billing price
            assessment_billing_list = BillingList.objects.filter(category='ASSESSMENT')
            total_acad_term_billings = AcadTermBilling.objects.filter(
                billing__in=assessment_billing_list,
                year_level=student_data['year_level'],
                semester=student_data['semester']
            ).aggregate(total_price=Sum('price'))['total_price'] or 0

            # Fetch receipts
            receipts = Receipt.objects.filter(student=student, school_year=student_data['academic_year'])
            receipts_data = ReceiptSerializer(receipts, many=True).data

            # Generate the COR file
            file_path = self.generate_registration_form(student_data, enrollments_data, joined_data, total_acad_term_billings, receipts_data)

            # Return the file as a downloadable response
            return FileResponse(
                open(file_path, 'rb'), as_attachment=True, filename=f"registration_form_{student_id}.xlsx"
            )
        except Student.DoesNotExist:
            return Response({"error": "Student not found."}, status=404)
        except Exception as e:
            return Response({"error": str(e)}, status=500)

    @staticmethod
    def generate_registration_form(student_data, enrollments_data, joined_data, total_acad_term_billings, receipts_data):
    # Create a workbook and sheet
        wb = Workbook()
        sheet = wb.active
        sheet.title = "Registration Form"

        # Add logo image
        logo_path = os.path.join(settings.BASE_DIR, 'static/images/Uni_Logo.png')  # UNIVERSITY LOGO
        if os.path.exists(logo_path):
            img = Image(logo_path)
            img.width, img.height = 150, 75
            sheet.add_image(img, "A1")

        # Add header information
        header_font = Font(size=12, bold=True)
        sheet.merge_cells("A2:H2")
        sheet["A2"] = "Republic of the Philippines"
        sheet["A2"].font = header_font
        
        sheet.merge_cells("A3:H3")
        sheet["A3"] = "CAVITE STATE UNIVERSITY"
        sheet["A3"].font = Font(size=14, bold=True)

        sheet.merge_cells("A4:H4")
        sheet["A4"] = "Bacoor Campus"
        sheet["A4"].font = header_font

        sheet.merge_cells("A6:H6")
        sheet["A6"] = "REGISTRATION FORM"
        sheet["A6"].font = Font(size=16, bold=True)

        # Add student information
        sheet["A9"] = "Student Number:"
        sheet["B9"] = student_data['id']

        sheet["A10"] = "Student Name:"
        sheet["B10"] = f"{student_data['last_name']}, {student_data['first_name']} {student_data.get('middle_name', '')}"

        sheet["A11"] = "Course:"
        sheet["B11"] = student_data['program']
        sheet["C11"] = "Year:"
        sheet["D11"] = student_data['year_level']

        sheet["A12"] = "Address:"
        address = student_data.get('address', {})
        sheet["B12"] = f"{address.get('street', '')}, {address.get('barangay', '')}, {address.get('city', '')}, {address.get('province', '')}"

        # Add course details header
        sheet["A14"] = "Course Code"
        sheet["B14"] = "Course Title"
        sheet["C14"] = "Units"
        sheet["D14"] = "Time"
        sheet["E14"] = "Day"
        sheet["F14"] = "Room"
        for col in ["A14", "B14", "C14", "D14", "E14", "F14"]:
            sheet[col].font = Font(bold=True)

        # Add course details
        row = 15
        for enrollment in enrollments_data:
            sheet[f"A{row}"] = enrollment['course']['code']
            sheet[f"B{row}"] = enrollment['course']['title']
            sheet[f"C{row}"] = enrollment['course']['lab_units'] + enrollment['course']['lec_units']
            sheet[f"D{row}"] = "TBA"
            sheet[f"E{row}"] = "TBA"
            sheet[f"F{row}"] = "TBA"
            row += 1

        # Add billing details header
        row += 2
        sheet[f"A{row}"] = "Lab Fees"
        sheet[f"B{row}"] = "Other Fees"
        sheet[f"C{row}"] = "Assessment"
        sheet[f"D{row}"] = "Payments"
        for col in [f"A{row}", f"B{row}", f"C{row}", f"D{row}"]:
            sheet[col].font = Font(bold=True)

        # Add billing data
        row += 1
        for billing in joined_data:
            sheet[f"A{row}"] = billing['billing_list']['name']
            sheet[f"B{row}"] = billing['billing_list']['price']
            row += 1

        # Add total billing price
        sheet[f"A{row}"] = "Total Billing Price"
        sheet[f"B{row}"] = total_acad_term_billings

        # Save the file
        file_path = os.path.join(settings.MEDIA_ROOT, f"registration_form_{student_data['id']}.xlsx")
        wb.save(file_path)
        return file_path

    
# Importing
class ImportExcelView(APIView):
    parser_classes = [MultiPartParser]# Specify the parser to handle multipart file uploads
    # Validates that the required columns are present in the uploaded Excel file.
    def validate_columns(self, data, required_columns): 
        missing_columns = set(required_columns) - set(data.columns)
        if missing_columns:
            raise ValueError(f"Missing required columns: {missing_columns}")#Raises a ValueError if any required column is missing.
    
    # Processes the student data from the Excel file.
    def process_students(self, data):
        required_columns = [
            "id", "first_name", "last_name", "email",
            "contact_number", "program", "gender", "year_level", "status"
        ]
        self.validate_columns(data, required_columns)# Validate that all required columns are present in the data


        students = [
            Student(
                id=row["id"],
                first_name=row["first_name"],
                last_name=row["last_name"],
                email=row["email"],
                contact_number=row["contact_number"],
                program_id=row["program"],
                gender=row["gender"],
                year_level=row["year_level"],
                status=row["status"],
            )
            for _, row in data.iterrows()
        ]
        with transaction.atomic():  # Perform a bulk insert operation within a database transaction
            Student.objects.bulk_create(students, ignore_conflicts=True)
        return f"Successfully imported {len(students)} students."

    # Processes the grade data from the Excel file.
    def process_grades(self, data):
        required_columns = ["student_id", "course_code", "grade", "semester", "academic_year"]
        self.validate_columns(data, required_columns)

        grades = [
            Grade(
                student_id=row["student_id"],
                course_code=row["course_code"],
                grade=row["grade"],
                semester=row["semester"],
                academic_year=row["academic_year"],
            )
            for _, row in data.iterrows()
        ]
        with transaction.atomic():
            Grade.objects.bulk_create(grades, ignore_conflicts=True)
        return f"Successfully imported {len(grades)} grades."

    # Handles the POST request to upload an Excel file.
    def post(self, request, *args, **kwargs):
        file = request.FILES.get('file', None) # Retrieve the uploaded file and the type of data being processed
        data_type = request.data.get('type', None)

        # Check if a file was uploaded
        if not file:
            return Response({"error": "No file uploaded."}, status=400)
        if data_type not in ['students', 'grades']:
            return Response({"error": "Invalid or missing data type. Use 'students' or 'grades'."}, status=400)

        try: # Read the uploaded Excel file into a Pandas DataFrame
            data = pd.read_excel(file)

            if data_type == 'students':
                message = self.process_students(data)
            elif data_type == 'grades':
                message = self.process_grades(data)

            return Response({"message": message}, status=200) # Return a success response with the result

        except ValueError as e:
            return Response({"error": str(e)}, status=400) # Return a response for validation errors

        except Exception as e:
            return Response({"error": f"An unexpected error occurred: {str(e)}"}, status=500)# Handle unexpected errors
